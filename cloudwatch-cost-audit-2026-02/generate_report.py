#!/usr/bin/env python3
"""Generate CloudWatch Cost Optimization Audit Report for Luckin Coffee North America."""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# DATA: All metric namespaces and counts from the audit
# ============================================================================
NAMESPACE_DATA = {
    "AWS/EMRServerless": 895562,
    "AWS/ElastiCache": 17359,
    "AWS/Kafka": 10891,
    "AWS/EBS": 4985,
    "AWS/EC2": 4256,
    "AWS/ES": 2011,
    "AWS/RDS": 1769,
    "AWS/DocDB": 1097,
    "AWS/NetworkELB": 953,
    "AWS/Usage": 812,
    "AWS/TrustedAdvisor": 489,
    "AWS/Logs": 194,
    "AWS/TransitGateway": 144,
    "AWS/InternetMonitor": 130,
    "AWS/CloudFront": 118,
    "AWS/EKS": 116,
    "AWS/WorkMail": 104,
    "AWS/Lambda": 94,
    "AWS/Route53": 76,
    "AWS/PrivateLinkEndpoints": 75,
    "AWS/SQS": 68,
    "AWS/S3": 68,
    "AWS/AutoScaling": 63,
    "AWS/NATGateway": 62,
    "AWS/EFS": 62,
    "AWS/ApplicationELB": 58,
    "AWS/ECR": 49,
    "AWS/ElasticMapReduce": 40,
    "AWS/Events": 34,
    "AWS/Backup": 20,
    "AWS/SNS": 14,
    "AWS/WAFV2": 13,
    "AWS/SES": 12,
    "AWS/ApiGateway": 11,
    "AWS/Firehose": 9,
    "AWS/GuardDuty": 8,
    "AWSLicenseManager/licenseUsage": 7,
    "AWS/HealthLake": 6,
    "AWS/Polly": 5,
    "AWS/KMS": 5,
    "AWS/CertificateManager": 5,
    "AWS/Bedrock/DataAutomation": 5,
    "AWS/SSM-RunCommand": 3,
    "AWS/X-Ray": 1,
    "AWS/SecretsManager": 1,
    "AWS/IPAM": 1,
}

TOTAL_METRICS = sum(NAMESPACE_DATA.values())

# System category mapping
CATEGORY_MAP = {
    "AWS/EMRServerless": "Data Processing (EMR Serverless)",
    "AWS/ElastiCache": "Redis Cache Clusters",
    "AWS/Kafka": "Kafka/MSK Messaging",
    "AWS/EBS": "Block Storage (EBS)",
    "AWS/EC2": "EC2 Compute Instances",
    "AWS/ES": "OpenSearch/Elasticsearch",
    "AWS/RDS": "MySQL/PostgreSQL RDS",
    "AWS/DocDB": "DocumentDB (MongoDB-compatible)",
    "AWS/NetworkELB": "Network Load Balancers",
    "AWS/Usage": "AWS Usage Tracking",
    "AWS/TrustedAdvisor": "Trusted Advisor",
    "AWS/Logs": "CloudWatch Logs",
    "AWS/TransitGateway": "Transit Gateway Networking",
    "AWS/InternetMonitor": "Internet Monitor",
    "AWS/CloudFront": "CDN (CloudFront)",
    "AWS/EKS": "Kubernetes (EKS)",
    "AWS/WorkMail": "WorkMail",
    "AWS/Lambda": "Serverless Functions",
    "AWS/Route53": "DNS (Route 53)",
    "AWS/PrivateLinkEndpoints": "PrivateLink VPC Endpoints",
    "AWS/SQS": "Message Queues (SQS)",
    "AWS/S3": "Object Storage (S3)",
    "AWS/AutoScaling": "Auto Scaling Groups",
    "AWS/NATGateway": "NAT Gateways",
    "AWS/EFS": "Elastic File System",
    "AWS/ApplicationELB": "Application Load Balancers",
    "AWS/ECR": "Container Registry (ECR)",
    "AWS/ElasticMapReduce": "EMR (Hadoop)",
    "AWS/Events": "EventBridge",
    "AWS/Backup": "AWS Backup",
    "AWS/SNS": "Notifications (SNS)",
    "AWS/WAFV2": "Web Application Firewall",
    "AWS/SES": "Email (SES)",
    "AWS/ApiGateway": "API Gateway",
    "AWS/Firehose": "Kinesis Firehose",
    "AWS/GuardDuty": "Security (GuardDuty)",
    "AWSLicenseManager/licenseUsage": "License Manager",
    "AWS/HealthLake": "HealthLake",
    "AWS/Polly": "Text-to-Speech (Polly)",
    "AWS/KMS": "Key Management (KMS)",
    "AWS/CertificateManager": "Certificate Manager",
    "AWS/Bedrock/DataAutomation": "Bedrock Data Automation",
    "AWS/SSM-RunCommand": "Systems Manager",
    "AWS/X-Ray": "X-Ray Tracing",
    "AWS/SecretsManager": "Secrets Manager",
    "AWS/IPAM": "IP Address Management",
}

# Detailed metric breakdown for top namespaces
DETAILED_METRICS = [
    # EMR Serverless - THE BIG ONE
    ("AWS/EMRServerless", "WorkerCpuAllocated", 111986, "Per-worker CPU allocation", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Per-worker metrics with ~112K unique dimensions from historical job runs. These are retained metrics from completed jobs."),
    ("AWS/EMRServerless", "WorkerEphemeralStorageAllocated", 111982, "Per-worker storage allocation", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "WorkerStorageWriteBytes", 111982, "Per-worker storage writes", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "WorkerEphemeralStorageUsed", 111981, "Per-worker storage usage", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "WorkerMemoryAllocated", 111980, "Per-worker memory allocation", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "WorkerMemoryUsed", 111980, "Per-worker memory usage", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "WorkerCpuUsed", 111975, "Per-worker CPU usage", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "WorkerStorageReadBytes", 111974, "Per-worker storage reads", "Data Processing (EMR Serverless)", "Remove", "HIGH", "Same issue - historical per-worker dimensions"),
    ("AWS/EMRServerless", "CPUAllocated (agg)", 6, "Aggregate CPU allocated", "Data Processing (EMR Serverless)", "Keep", "LOW", "Aggregate job-level metrics are fine"),
    ("AWS/EMRServerless", "Other agg metrics (6 types)", 30, "Job-level aggregate metrics", "Data Processing (EMR Serverless)", "Keep", "LOW", "Aggregate metrics - keep"),

    # ElastiCache - 156 clusters
    ("AWS/ElastiCache", "Standard metrics (42 types)", 13146, "ActiveDefragHits, CPUUtilization, CacheHits, etc.", "Redis Cache Clusters", "Keep", "LOW", "Standard per-node metrics for 156 clusters. AWS-published, no extra cost."),
    ("AWS/ElastiCache", "Command-type metrics (HashBased, SetBased, etc.)", 2607, "Per-command-type latency and count", "Redis Cache Clusters", "Review", "MEDIUM", "Command-specific metrics - review if all are monitored. Many may be unused."),
    ("AWS/ElastiCache", "CPU Credit metrics", 618, "CPUCreditBalance, CPUCreditUsage", "Redis Cache Clusters", "Keep", "LOW", "Important for burstable t4g instances"),
    ("AWS/ElastiCache", "Other specialized metrics", 988, "CacheHitRate, ErrorCount, etc.", "Redis Cache Clusters", "Keep", "LOW", "Standard monitoring metrics"),

    # Kafka/MSK - 3 clusters
    ("AWS/Kafka", "OffsetLag", 1920, "Per-topic, per-consumer-group lag", "Kafka/MSK Messaging", "Review", "MEDIUM", "High cardinality - 1920 dimensions across 3 clusters. Review consumer groups."),
    ("AWS/Kafka", "EstimatedTimeLag", 1918, "Consumer time lag estimation", "Kafka/MSK Messaging", "Review", "MEDIUM", "Same high cardinality as OffsetLag"),
    ("AWS/Kafka", "BytesOutPerSec", 876, "Per-topic byte output rate", "Kafka/MSK Messaging", "Keep", "LOW", "Important for monitoring throughput"),
    ("AWS/Kafka", "BytesInPerSec/MessagesInPerSec", 1088, "Per-topic input metrics", "Kafka/MSK Messaging", "Keep", "LOW", "Important for monitoring throughput"),
    ("AWS/Kafka", "Per-broker metrics (80+ types)", 5089, "CPU, memory, disk, network per broker", "Kafka/MSK Messaging", "Keep", "LOW", "Standard broker-level metrics"),

    # EBS
    ("AWS/EBS", "Volume metrics (10 types)", 4985, "ReadOps, WriteOps, QueueLength, etc.", "Block Storage (EBS)", "Keep", "LOW", "Standard per-volume metrics for ~332 volumes"),

    # EC2
    ("AWS/EC2", "Standard metrics (18 types)", 4256, "CPU, Network, EBS, Status checks", "EC2 Compute Instances", "Keep", "LOW", "Standard per-instance metrics for 233 instances"),

    # OpenSearch
    ("AWS/ES", "Per-node metrics (52+ types)", 1011, "CPU, Memory, JVM, I/O per node", "OpenSearch/Elasticsearch", "Keep", "LOW", "Standard per-node metrics for 4 domains"),
    ("AWS/ES", "Cluster-level metrics (100+ types)", 1000, "Cluster health, shards, search, indexing", "OpenSearch/Elasticsearch", "Keep", "LOW", "Important for cluster monitoring"),

    # RDS
    ("AWS/RDS", "Standard metrics (16 types)", 1200, "CPU, Connections, IOPS, Latency", "MySQL/PostgreSQL RDS", "Keep", "LOW", "Standard per-instance metrics for 74 instances"),
    ("AWS/RDS", "MySQL-specific (BinLog, LVM)", 207, "BinLogDiskUsage, LVMReadIOPS", "MySQL/PostgreSQL RDS", "Keep", "LOW", "MySQL-specific metrics"),
    ("AWS/RDS", "PostgreSQL-specific", 49, "CheckpointLag, TransactionLogs", "MySQL/PostgreSQL RDS", "Keep", "LOW", "PostgreSQL-specific metrics"),
    ("AWS/RDS", "Performance Insights", 23, "DBLoad, DBLoadCPU, DBLoadNonCPU", "MySQL/PostgreSQL RDS", "Review", "MEDIUM", "PI costs extra - verify it's enabled only where needed"),

    # DocumentDB
    ("AWS/DocDB", "Standard metrics (54+ types)", 1097, "CPU, Connections, IOPS, Opcounters", "DocumentDB (MongoDB-compatible)", "Keep", "LOW", "Standard metrics for 3 clusters (devops, gia, iot)"),

    # Network/Load Balancers
    ("AWS/NetworkELB", "NLB metrics", 953, "HealthyHostCount, ProcessedBytes, etc.", "Network Load Balancers", "Keep", "LOW", "Standard NLB metrics"),
    ("AWS/ApplicationELB", "ALB metrics", 58, "RequestCount, ResponseTime, etc.", "Application Load Balancers", "Keep", "LOW", "Standard ALB metrics"),

    # CloudWatch Logs
    ("AWS/Logs", "IncomingLogEvents/Bytes", 184, "Log ingestion metrics", "CloudWatch Logs", "Keep", "LOW", "Log volume tracking"),

    # InternetMonitor
    ("AWS/InternetMonitor", "All metrics", 130, "PerformanceScore, Availability, RoundTripTime", "Internet Monitor", "Review", "MEDIUM", "At $37/month for monitored resources - evaluate if needed"),

    # CloudFront
    ("AWS/CloudFront", "CDN metrics", 118, "Requests, ErrorRate, BytesDownloaded", "CDN (CloudFront)", "Keep", "LOW", "Standard CDN monitoring"),

    # EKS
    ("AWS/EKS", "Flow control metrics", 116, "apiserver_flowcontrol_* per priority level", "Kubernetes (EKS)", "Keep", "LOW", "Standard EKS control plane metrics"),

    # WorkMail
    ("AWS/WorkMail", "ActionDenied", 79, "Denied action events", "WorkMail", "Review", "LOW", "High count of ActionDenied - investigate"),
    ("AWS/WorkMail", "Other metrics", 25, "AuthFailure, AccessDenied, Email", "WorkMail", "Keep", "LOW", "Standard email metrics"),

    # Lambda
    ("AWS/Lambda", "Function metrics", 94, "Invocations, Errors, Duration, Throttles", "Serverless Functions", "Keep", "LOW", "Standard function metrics for ~13 functions"),

    # Other small namespaces
    ("AWS/Route53", "Health checks + DNS", 76, "HealthCheckStatus, TTFByte, DNSQueries", "DNS (Route 53)", "Keep", "LOW", "Standard Route 53 metrics"),
    ("AWS/PrivateLinkEndpoints", "VPC endpoint metrics", 75, "Connections, Packets, Bytes", "PrivateLink VPC Endpoints", "Keep", "LOW", "Standard endpoint metrics"),
    ("AWS/SQS", "Queue metrics", 68, "MessagesSent/Received, QueueDepth", "Message Queues (SQS)", "Keep", "LOW", "Standard queue metrics"),
    ("AWS/S3", "Bucket metrics", 68, "BucketSizeBytes, NumberOfObjects", "Object Storage (S3)", "Keep", "LOW", "Standard bucket metrics"),
    ("AWS/NATGateway", "NAT metrics", 62, "BytesIn/Out, Connections, Packets", "NAT Gateways", "Keep", "LOW", "Standard NAT metrics"),
    ("AWS/EFS", "EFS metrics", 62, "StorageBytes, IOBytes, ClientConnections", "Elastic File System", "Keep", "LOW", "Standard EFS metrics"),
    ("AWS/AutoScaling", "ASG metrics", 63, "GroupCapacity, WarmPool metrics", "Auto Scaling Groups", "Keep", "LOW", "Standard ASG metrics"),
    ("AWS/ECR", "RepositoryPullCount", 49, "Pull count per repo", "Container Registry (ECR)", "Keep", "LOW", "Standard ECR metrics"),
    ("AWS/ElasticMapReduce", "EMR metrics", 40, "HDFS, YARN, S3 IO", "EMR (Hadoop)", "Review", "LOW", "2 alarms in INSUFFICIENT_DATA - check if EMR clusters still exist"),
    ("AWS/Events", "EventBridge metrics", 34, "MatchedEvents, Invocations", "EventBridge", "Keep", "LOW", "Standard EventBridge metrics"),
    ("AWS/Backup", "Backup metrics", 20, "RecoveryPoints, BackupJobs", "AWS Backup", "Keep", "LOW", "Standard backup metrics"),
    ("AWS/TrustedAdvisor", "TA checks", 489, "Trusted Advisor check metrics", "Trusted Advisor", "Keep", "LOW", "AWS-published, no extra cost"),
    ("AWS/Usage", "API usage tracking", 812, "CallCount, ResourceCount per service", "AWS Usage Tracking", "Keep", "LOW", "AWS-published usage metrics"),
    ("AWS/TransitGateway", "TGW metrics", 144, "Bytes/Packets In/Out per attachment", "Transit Gateway Networking", "Keep", "LOW", "Standard TGW metrics"),
]

# Cost data (January 2026)
COST_DATA = {
    "CW:MetricMonitorUsage": {"desc": "Custom Metrics Storage", "jan": 1872.68, "dec": 1891.83, "nov": 1920.66},
    "CW:Requests": {"desc": "API Requests (GetMetricData, PutMetricData)", "jan": 1507.72, "dec": 1503.68, "nov": 982.31},
    "CW:GMD-Metrics": {"desc": "GetMetricData Metrics Scanned", "jan": 231.91, "dec": 3.60, "nov": 3.46},
    "InternetMonitor-MonitoredResource": {"desc": "Internet Monitor Resources", "jan": 37.36, "dec": 37.33, "nov": 36.09},
    "InternetMonitor-CityNetwork": {"desc": "Internet Monitor City Networks", "jan": 6.61, "dec": 5.83, "nov": 6.71},
    "DataProcessing-Bytes": {"desc": "Logs Data Processing", "jan": 4.15, "dec": 2.75, "nov": 3.41},
    "DashboardsUsageHour": {"desc": "Dashboard Hours", "jan": 3.00, "dec": 3.00, "nov": 4.91},
    "AlarmMonitorUsage": {"desc": "Alarm Monitoring", "jan": 1.74, "dec": 1.19, "nov": 1.05},
    "MetricInsightAlarmUsage": {"desc": "Metric Insights Alarms", "jan": 1.40, "dec": 1.40, "nov": 1.40},
    "TimedStorage-ByteHrs": {"desc": "Logs Storage", "jan": 0.21, "dec": 0.16, "nov": 0.11},
}

TOTAL_JAN_COST = sum(v["jan"] for v in COST_DATA.values())

# Alarm summary
ALARM_SUMMARY = [
    ("AWS ES disk space alert (ES)", "AWS/ES (composite)", "OK", True, "Keep"),
    ("EMR MRLostNodes", "AWS/ElasticMapReduce", "INSUFFICIENT_DATA", True, "Review - EMR may be decommissioned"),
    ("HDFSUtilization", "AWS/ElasticMapReduce", "INSUFFICIENT_DATA", True, "Review - EMR may be decommissioned"),
    ("Kafka partition/disk/offline (9 alarms)", "AWS/Kafka", "OK", True, "Keep - critical MSK monitoring"),
    ("API Gateway img-transform (3 alarms)", "AWS/ApiGateway", "OK", True, "Keep - image transform monitoring"),
    ("CDN img-transform (3 alarms)", "AWS/CloudFront", "OK", True, "Keep - CDN monitoring"),
    ("Lambda img-transform (3 alarms)", "AWS/Lambda", "OK", True, "Keep - Lambda monitoring"),
    ("CDN domain 5xx alerts (7 alarms)", "AWS/CloudFront", "OK/INSUF", True, "Review - 2 in INSUFFICIENT_DATA"),
]


def create_xlsx_report():
    wb = Workbook()

    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    remove_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    remove_font = Font(name="Calibri", bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=10)
    money_format = '#,##0.00'
    number_format = '#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, wrap=False):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        return cell

    # ========================================================================
    # Sheet 1: Executive Summary
    # ========================================================================
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.cell(row=1, column=1, value="CloudWatch Cost Optimization Audit").font = title_font
    ws.cell(row=2, column=1, value="Luckin Coffee North America - AWS Account").font = bold_font
    ws.cell(row=3, column=1, value=f"Audit Date: {datetime.now().strftime('%Y-%m-%d')}").font = normal_font
    ws.cell(row=4, column=1, value=f"Report Period: November 2025 - February 2026").font = normal_font

    # Key Metrics
    r = 6
    ws.cell(row=r, column=1, value="KEY FINDINGS").font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    r += 1
    summary_items = [
        ("Total Active Metric Dimensions", f"{TOTAL_METRICS:,}"),
        ("Total Namespaces", f"{len(NAMESPACE_DATA)}"),
        ("January 2026 CloudWatch Cost", f"${TOTAL_JAN_COST:,.2f}"),
        ("Annualized Cost (at Jan rate)", f"${TOTAL_JAN_COST * 12:,.2f}"),
        ("", ""),
        ("COST BREAKDOWN (Jan 2026)", ""),
        ("  Custom Metrics Storage", f"${COST_DATA['CW:MetricMonitorUsage']['jan']:,.2f} (51.1%)"),
        ("  API Requests", f"${COST_DATA['CW:Requests']['jan']:,.2f} (41.1%)"),
        ("  GetMetricData Scanned", f"${COST_DATA['CW:GMD-Metrics']['jan']:,.2f} (6.3%) -- 64x SPIKE from Dec!"),
        ("  Internet Monitor", f"${COST_DATA['InternetMonitor-MonitoredResource']['jan'] + COST_DATA['InternetMonitor-CityNetwork']['jan']:,.2f} (1.2%)"),
        ("  Other (Dashboards, Alarms, Logs)", "$10.50 (0.3%)"),
        ("", ""),
        ("BIGGEST COST DRIVER", ""),
        ("  AWS/EMRServerless", f"{NAMESPACE_DATA['AWS/EMRServerless']:,} metric dimensions (95.2% of all metrics!)"),
        ("  - 8 per-worker metrics x ~112K unique worker dimensions", ""),
        ("  - These are retained from historical job runs", ""),
        ("  - EMR Serverless auto-generates per-worker metrics that persist", ""),
        ("", ""),
        ("ESTIMATED SAVINGS OPPORTUNITY", ""),
        ("  EMR Serverless worker metric cleanup", "$800-1,200/month"),
        ("  GetMetricData query optimization", "$200-230/month"),
        ("  Internet Monitor review", "$37-44/month"),
        ("  Unused alarm cleanup", "$2-5/month"),
        ("  Total Potential Savings", "$1,039-1,479/month ($12,468-17,748/year)"),
        ("", ""),
        ("INFRASTRUCTURE SUMMARY", ""),
        ("  EC2 Instances (running)", "233"),
        ("  RDS Instances (MySQL/PostgreSQL)", "65 (62 MySQL + 3 PostgreSQL)"),
        ("  DocumentDB Instances", "9 (3 clusters: devops, gia, iot)"),
        ("  ElastiCache Clusters", "156 (Redis)"),
        ("  MSK/Kafka Clusters", "3 (architecture, base, business)"),
        ("  OpenSearch Domains", "4 (luckycommon, luckylfe-log, luckyur-log, luckyus-opensearch-dify)"),
        ("  EBS Volumes", "~332"),
        ("  CloudWatch Alarms", "28 (25 OK, 3 INSUFFICIENT_DATA)"),
        ("  EC2 Detailed Monitoring", "8 instances (225 basic)"),
    ]
    for item in summary_items:
        ws.cell(row=r, column=1, value=item[0]).font = bold_font if item[0] and not item[0].startswith("  ") else normal_font
        ws.cell(row=r, column=2, value=item[1]).font = normal_font
        r += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 55

    # ========================================================================
    # Sheet 2: Namespace Overview
    # ========================================================================
    ws2 = wb.create_sheet("Namespace Overview")
    ws2.sheet_properties.tabColor = "2E75B6"

    headers = ["Namespace", "Metric Count", "% of Total", "System Category", "Type", "Estimated Monthly Cost Contribution"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header_row(ws2, 1, len(headers))

    row = 2
    for ns, count in sorted(NAMESPACE_DATA.items(), key=lambda x: -x[1]):
        pct = count / TOTAL_METRICS * 100
        cat = CATEGORY_MAP.get(ns, "Other")
        ns_type = "AWS Native" if ns.startswith("AWS/") else "Custom/3rd Party"

        # Rough cost estimation
        if ns == "AWS/EMRServerless":
            est_cost = "$1,200-1,500 (dominant driver)"
        elif count > 10000:
            est_cost = f"${count * 0.10 / 1000:.0f}-{count * 0.30 / 1000:.0f}"
        elif count > 1000:
            est_cost = f"${count * 0.05 / 1000:.0f}-{count * 0.10 / 1000:.0f}"
        else:
            est_cost = "< $1 (within free tier or standard)"

        c = style_data_cell(ws2, row, 1)
        c.value = ns
        style_data_cell(ws2, row, 2).value = count
        ws2.cell(row=row, column=2).number_format = number_format
        style_data_cell(ws2, row, 3).value = round(pct, 2)
        ws2.cell(row=row, column=3).number_format = '0.00"%"'
        style_data_cell(ws2, row, 4).value = cat
        style_data_cell(ws2, row, 5).value = ns_type
        style_data_cell(ws2, row, 6).value = est_cost

        # Color-code by size
        if count > 100000:
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = high_fill
        elif count > 5000:
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = medium_fill

        row += 1

    # Totals row
    ws2.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws2.cell(row=row, column=2, value=TOTAL_METRICS).font = bold_font
    ws2.cell(row=row, column=2).number_format = number_format
    ws2.cell(row=row, column=3, value=100.0).font = bold_font
    ws2.cell(row=row, column=6, value=f"~${TOTAL_JAN_COST:,.2f}/month actual").font = bold_font

    for col in [1, 2, 3, 4, 5, 6]:
        ws2.column_dimensions[get_column_letter(col)].width = [35, 15, 12, 35, 18, 35][col-1]

    # ========================================================================
    # Sheet 3: Detailed Metric Inventory
    # ========================================================================
    ws3 = wb.create_sheet("Detailed Metric Inventory")
    ws3.sheet_properties.tabColor = "548235"

    headers3 = ["Namespace", "MetricName / Group", "Dimension Count", "Description", "System Category",
                 "Recommendation", "Priority", "Notes"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header_row(ws3, 1, len(headers3))

    row = 2
    for item in DETAILED_METRICS:
        ns, metric, dims, desc, cat, rec, priority, notes = item
        for col, val in enumerate([ns, metric, dims, desc, cat, rec, priority, notes], 1):
            c = style_data_cell(ws3, row, col, wrap=(col in [4, 8]))
            c.value = val
            if col == 3 and isinstance(val, int):
                c.number_format = number_format

        # Color code recommendation
        rec_cell = ws3.cell(row=row, column=6)
        if rec == "Remove":
            rec_cell.fill = remove_fill
            rec_cell.font = remove_font
        elif rec == "Review":
            rec_cell.fill = medium_fill

        # Color code priority
        pri_cell = ws3.cell(row=row, column=7)
        if priority == "HIGH":
            pri_cell.fill = high_fill
        elif priority == "MEDIUM":
            pri_cell.fill = medium_fill
        else:
            pri_cell.fill = low_fill

        row += 1

    for col, w in enumerate([30, 35, 15, 40, 30, 15, 12, 55], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ========================================================================
    # Sheet 4: Cost Trend Analysis
    # ========================================================================
    ws4 = wb.create_sheet("Cost Trend Analysis")
    ws4.sheet_properties.tabColor = "BF8F00"

    headers4 = ["Usage Type", "Description", "Nov 2025", "Dec 2025", "Jan 2026", "Feb 2026 (projected)", "MoM Change", "Trend"]
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=i, value=h)
    style_header_row(ws4, 1, len(headers4))

    row = 2
    for key, data in COST_DATA.items():
        feb_proj = data["jan"]  # simple projection
        mom_change = ((data["jan"] - data["dec"]) / data["dec"] * 100) if data["dec"] > 0 else 0
        trend = "UP" if mom_change > 5 else ("DOWN" if mom_change < -5 else "STABLE")

        vals = [key, data["desc"], data["nov"], data["dec"], data["jan"], feb_proj, f"{mom_change:+.1f}%", trend]
        for col, val in enumerate(vals, 1):
            c = style_data_cell(ws4, row, col)
            c.value = val
            if col in [3, 4, 5, 6] and isinstance(val, (int, float)):
                c.number_format = money_format

        # Highlight significant increases
        if mom_change > 50:
            for col in range(1, 9):
                ws4.cell(row=row, column=col).fill = high_fill
        elif mom_change > 10:
            for col in range(1, 9):
                ws4.cell(row=row, column=col).fill = medium_fill

        row += 1

    # Total row
    row += 1
    ws4.cell(row=row, column=1, value="TOTAL").font = bold_font
    for month_col, month_key in [(3, "nov"), (4, "dec"), (5, "jan")]:
        total = sum(v[month_key] for v in COST_DATA.values())
        ws4.cell(row=row, column=month_col, value=total).font = bold_font
        ws4.cell(row=row, column=month_col).number_format = money_format

    for col, w in enumerate([35, 35, 14, 14, 14, 18, 14, 10], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # ========================================================================
    # Sheet 5: Alarms & Dashboards
    # ========================================================================
    ws5 = wb.create_sheet("Alarms & Dashboards")
    ws5.sheet_properties.tabColor = "C55A11"

    headers5 = ["Alarm Name/Group", "Namespace", "State", "Actions Enabled", "Recommendation"]
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=i, value=h)
    style_header_row(ws5, 1, len(headers5))

    row = 2
    for alarm in ALARM_SUMMARY:
        for col, val in enumerate(alarm, 1):
            c = style_data_cell(ws5, row, col, wrap=(col == 5))
            c.value = str(val)
        if "INSUFFICIENT" in alarm[2]:
            ws5.cell(row=row, column=3).fill = medium_fill
        row += 1

    row += 2
    ws5.cell(row=row, column=1, value="DASHBOARD SUMMARY").font = bold_font
    row += 1
    ws5.cell(row=row, column=1, value="No CloudWatch dashboards found in this account.").font = normal_font
    row += 1
    ws5.cell(row=row, column=1, value="Note: Dashboards cost $3/month each. Currently $3/month in dashboard charges").font = normal_font
    ws5.cell(row=row + 1, column=1, value="suggests 1 dashboard exists (possibly in another region or from a linked feature).").font = normal_font

    for col, w in enumerate([45, 25, 18, 18, 45], 1):
        ws5.column_dimensions[get_column_letter(col)].width = w

    # ========================================================================
    # Sheet 6: EC2 Detailed Monitoring
    # ========================================================================
    ws6 = wb.create_sheet("EC2 Monitoring Status")
    ws6.sheet_properties.tabColor = "7030A0"

    headers6 = ["Instance ID", "Instance Name", "Instance Type", "Monitoring State", "Recommendation"]
    for i, h in enumerate(headers6, 1):
        ws6.cell(row=1, column=i, value=h)
    style_header_row(ws6, 1, len(headers6))

    detailed_instances = [
        ("i-0d9b1e751b082509e", "unnamed", "m6a.xlarge", "enabled", "Review - unnamed instance with detailed monitoring"),
        ("i-0b7168740da06034a", "unnamed", "m6a.large", "enabled", "Review - unnamed instance with detailed monitoring"),
        ("i-031550c05c4eae3e9", "unnamed", "m6a.large", "enabled", "Review - unnamed instance with detailed monitoring"),
        ("i-0f5b8098e2098d6ec", "unnamed", "r6i.2xlarge", "enabled", "Review - unnamed instance with detailed monitoring"),
        ("i-0d812309d4071771c", "unnamed", "r6i.2xlarge", "enabled", "Review - unnamed instance with detailed monitoring"),
        ("i-0047137d18fc77c35", "velodb-agent-0dce7ca7770422d33", "c5.large", "enabled", "Review - VeloDB agent, may need detailed monitoring"),
        ("i-0a1e3602ff0d0cccf", "unnamed", "m6a.xlarge", "enabled", "Review - unnamed instance with detailed monitoring"),
        ("i-0418d2213e6688dc0", "unnamed", "m6a.large", "enabled", "Review - unnamed instance with detailed monitoring"),
    ]

    row = 2
    for inst in detailed_instances:
        for col, val in enumerate(inst, 1):
            c = style_data_cell(ws6, row, col, wrap=(col == 5))
            c.value = val
        ws6.cell(row=row, column=4).fill = medium_fill
        row += 1

    row += 1
    ws6.cell(row=row, column=1, value="Summary").font = bold_font
    ws6.cell(row=row + 1, column=1, value="Total instances: 233 running").font = normal_font
    ws6.cell(row=row + 2, column=1, value="Detailed monitoring enabled: 8 (3.4%)").font = normal_font
    ws6.cell(row=row + 3, column=1, value="Basic monitoring: 225 (96.6%)").font = normal_font
    ws6.cell(row=row + 4, column=1, value="Detailed monitoring adds ~$2.10/instance/month = ~$16.80/month for 8 instances").font = normal_font
    ws6.cell(row=row + 5, column=1, value="Most are unnamed - likely EKS nodes managed by Auto Scaling. Review if 1-minute metrics are needed.").font = normal_font

    for col, w in enumerate([25, 35, 15, 18, 50], 1):
        ws6.column_dimensions[get_column_letter(col)].width = w

    # ========================================================================
    # Sheet 7: Recommendations & Action Plan
    # ========================================================================
    ws7 = wb.create_sheet("Recommendations")
    ws7.sheet_properties.tabColor = "00B050"

    headers7 = ["Priority", "Category", "Recommendation", "Estimated Monthly Savings", "Effort", "Timeline"]
    for i, h in enumerate(headers7, 1):
        ws7.cell(row=1, column=i, value=h)
    style_header_row(ws7, 1, len(headers7))

    recommendations = [
        ("HIGH", "EMR Serverless", "Investigate 895K+ worker-level metrics from EMRServerless. These per-worker metrics persist from historical job runs and create extreme cardinality. Options: (1) Delete old EMR Serverless applications to clear retained metrics, (2) Configure metric filters to suppress per-worker metrics, (3) Use EMR Serverless application-level configs to reduce metric granularity.", "$800-1,200", "Medium", "1-2 weeks"),
        ("HIGH", "GetMetricData Spike", "CW:GMD-Metrics jumped from $3.60 to $231.91 (64x increase) in January 2026. Identify the source: new Grafana dashboards, monitoring tools, or automated queries hitting CloudWatch APIs excessively. Check Grafana data source polling intervals.", "$200-230", "Low", "1-3 days"),
        ("MEDIUM", "Internet Monitor", "Internet Monitor costs $37-44/month for monitored resources + city networks. Evaluate if this is actively used for Luckin Coffee's CDN/app monitoring or can be disabled in favor of CloudFront native metrics + Route 53 health checks.", "$37-44", "Low", "1 day"),
        ("MEDIUM", "Kafka Consumer Lag Metrics", "OffsetLag has 1,920 dimension combinations across 3 MSK clusters. Review and clean up unused consumer groups to reduce metric cardinality. Stale consumer groups generate metrics but serve no purpose.", "$5-15", "Medium", "1 week"),
        ("MEDIUM", "EMR Classic Alarms", "2 alarms (MRLostNodes, HDFSUtilization) are in INSUFFICIENT_DATA state, suggesting EMR classic clusters may be decommissioned. Clean up these alarms.", "$0.20", "Low", "1 hour"),
        ("MEDIUM", "CloudFront Alarms", "2 CloudFront alarms (ionepiecesync, web.luckincdn.us) are in INSUFFICIENT_DATA - distributions may be inactive. Verify and clean up.", "$0.20", "Low", "1 hour"),
        ("LOW", "EC2 Detailed Monitoring", "8 instances have detailed monitoring enabled (mostly unnamed, likely EKS nodes). If 1-minute resolution isn't needed, switch to basic 5-minute monitoring.", "$16.80", "Low", "1 hour"),
        ("LOW", "ElastiCache Command Metrics", "Review per-command-type metrics (HashBasedCmds, SetBasedCmds, etc.) across 156 Redis clusters. Many command-specific latency metrics may not be monitored.", "$0-5", "Medium", "1 week"),
        ("LOW", "WorkMail ActionDenied", "79 ActionDenied metric dimensions in WorkMail namespace. Investigate if this indicates a security issue or misconfiguration.", "$0", "Low", "1 day"),
        ("LOW", "RDS Performance Insights", "Performance Insights enabled on some instances (DBLoad metrics visible for 6 instances). Verify PI is only enabled where needed - it has separate pricing.", "$0-10", "Low", "1 day"),
    ]

    row = 2
    for rec in recommendations:
        for col, val in enumerate(rec, 1):
            c = style_data_cell(ws7, row, col, wrap=(col == 3))
            c.value = val

        pri_cell = ws7.cell(row=row, column=1)
        if rec[0] == "HIGH":
            pri_cell.fill = high_fill
        elif rec[0] == "MEDIUM":
            pri_cell.fill = medium_fill
        else:
            pri_cell.fill = low_fill

        row += 1

    # Summary
    row += 2
    ws7.cell(row=row, column=1, value="SAVINGS SUMMARY").font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    row += 1
    ws7.cell(row=row, column=1, value="Quick Wins (< 1 week):").font = bold_font
    ws7.cell(row=row, column=2, value="$237-274/month savings").font = bold_font
    row += 1
    ws7.cell(row=row, column=1, value="Medium-term (1-2 weeks):").font = bold_font
    ws7.cell(row=row, column=2, value="$805-1,215/month additional").font = bold_font
    row += 1
    ws7.cell(row=row, column=1, value="Total Potential:").font = Font(name="Calibri", bold=True, size=12, color="00B050")
    ws7.cell(row=row, column=2, value="$1,042-1,489/month ($12,504-17,868/year)").font = Font(name="Calibri", bold=True, size=12, color="00B050")
    row += 1
    ws7.cell(row=row, column=1, value="Current Monthly Spend:").font = bold_font
    ws7.cell(row=row, column=2, value=f"${TOTAL_JAN_COST:,.2f}/month").font = bold_font
    row += 1
    ws7.cell(row=row, column=1, value="Optimized Monthly Spend:").font = bold_font
    ws7.cell(row=row, column=2, value=f"~${TOTAL_JAN_COST - 1250:,.2f}/month (estimated)").font = bold_font

    for col, w in enumerate([18, 25, 80, 22, 12, 14], 1):
        ws7.column_dimensions[get_column_letter(col)].width = w

    # ========================================================================
    # Sheet 8: Infrastructure Inventory
    # ========================================================================
    ws8 = wb.create_sheet("Infrastructure Inventory")
    ws8.sheet_properties.tabColor = "4472C4"

    headers8 = ["Category", "Resource Type", "Count", "Details"]
    for i, h in enumerate(headers8, 1):
        ws8.cell(row=1, column=i, value=h)
    style_header_row(ws8, 1, len(headers8))

    infra = [
        ("Compute", "EC2 Instances", 233, "All running. 225 basic monitoring, 8 detailed"),
        ("Compute", "Lambda Functions", 13, "Image transform functions + others"),
        ("Compute", "EKS Clusters", "2+", "Kubernetes clusters for application workloads"),
        ("Database", "RDS MySQL Instances", 62, "Primarily db.t4g.micro (41) and db.t4g.medium (17)"),
        ("Database", "RDS PostgreSQL Instances", 3, "Smaller PostgreSQL deployment"),
        ("Database", "DocumentDB Clusters", 3, "docdb-devops, docdb-gia, docdb-iot (9 instances total)"),
        ("Cache", "ElastiCache Redis Clusters", 156, "Mostly cache.t4g.micro (129), cache.t4g.small (15)"),
        ("Messaging", "MSK/Kafka Clusters", 3, "iprod-kafka-architecture, iprod-kafka-base, iprod-kafka-business"),
        ("Search", "OpenSearch Domains", 4, "luckycommon, luckylfe-log, luckyur-log, luckyus-opensearch-dify"),
        ("Storage", "EBS Volumes", "~332", "Attached to EC2 instances"),
        ("Storage", "S3 Buckets", 34, "Tracked for size and object count"),
        ("Storage", "EFS File Systems", 4, "Elastic file storage"),
        ("Networking", "NLB (Network)", "50+", "Network load balancers"),
        ("Networking", "ALB (Application)", 4, "Application load balancers"),
        ("Networking", "NAT Gateways", 4, "VPC NAT gateways"),
        ("Networking", "Transit Gateways", "9+", "Cross-VPC connectivity"),
        ("Networking", "VPC Endpoints (PrivateLink)", 5, "PrivateLink endpoints"),
        ("CDN", "CloudFront Distributions", 11, "CDN for Luckin web/mobile/API"),
        ("Data Processing", "EMR Serverless Applications", "2+", "Spark/Hive jobs (source of 895K metrics)"),
        ("Data Processing", "EMR Classic Clusters", "1-2", "Legacy EMR (alarms in INSUFFICIENT_DATA)"),
        ("Data Processing", "Kinesis Firehose", 1, "Data delivery stream"),
        ("Security", "WAFv2", 1, "Web application firewall"),
        ("Security", "GuardDuty", 1, "Threat detection"),
        ("Communication", "WorkMail", 1, "Email service"),
        ("Communication", "SES", 1, "Email sending"),
        ("Communication", "SNS Topics", 4, "Notification topics"),
        ("Queuing", "SQS Queues", 5, "Message queues"),
        ("DNS", "Route 53", "30+", "Health checks and DNS queries"),
        ("AI/ML", "Bedrock Data Automation", 1, "Data automation pipeline"),
    ]

    row = 2
    for item in infra:
        for col, val in enumerate(item, 1):
            c = style_data_cell(ws8, row, col, wrap=(col == 4))
            c.value = val if not isinstance(val, int) else val
            if col == 3 and isinstance(val, int):
                c.number_format = number_format
        row += 1

    for col, w in enumerate([18, 30, 12, 60], 1):
        ws8.column_dimensions[get_column_letter(col)].width = w

    # Save
    output_path = "/app/cloudwatch-cost-audit-2026-02/cloudwatch_cost_audit_luckin_2026-02.xlsx"
    wb.save(output_path)
    print(f"XLSX saved to {output_path}")
    return output_path


def create_markdown_report():
    """Generate the markdown version of the report."""
    report = f"""# CloudWatch Cost Optimization Audit Report
## Luckin Coffee North America - AWS Account
**Audit Date:** {datetime.now().strftime('%Y-%m-%d')}
**Report Period:** November 2025 - February 2026

---

## Executive Summary

| Metric | Value |
|--------|-------|
| Total Active Metric Dimensions | {TOTAL_METRICS:,} |
| Total Namespaces | {len(NAMESPACE_DATA)} |
| January 2026 CloudWatch Cost | ${TOTAL_JAN_COST:,.2f} |
| Annualized Cost | ${TOTAL_JAN_COST * 12:,.2f} |
| Estimated Savings Opportunity | $1,042-1,489/month ($12,504-17,868/year) |

### Cost Breakdown (January 2026)

| Usage Type | Description | Cost | % of Total |
|------------|-------------|------|-----------|
| CW:MetricMonitorUsage | Custom Metrics Storage | $1,872.68 | 51.1% |
| CW:Requests | API Requests (Get/PutMetricData) | $1,507.72 | 41.1% |
| CW:GMD-Metrics | GetMetricData Metrics Scanned | $231.91 | 6.3% |
| InternetMonitor | Monitored Resources + City Networks | $43.97 | 1.2% |
| Other | Dashboards, Alarms, Logs, Storage | $10.50 | 0.3% |
| **TOTAL** | | **${TOTAL_JAN_COST:,.2f}** | **100%** |

### Critical Alert: GetMetricData Cost Spike

> **CW:GMD-Metrics jumped from $3.60 (Dec) to $231.91 (Jan) -- a 64x increase!**
> This suggests a new monitoring tool, Grafana dashboard, or automated query is scanning far more metric data points than before. Immediate investigation recommended.

---

## Top 10 Costliest Metric Groups

| Rank | Namespace | Metric Count | % of Total | System Category | Issue |
|------|-----------|-------------|-----------|----------------|-------|
| 1 | AWS/EMRServerless | 895,562 | 95.2% | Data Processing | **Per-worker metrics from historical jobs -- MASSIVE cardinality** |
| 2 | AWS/ElastiCache | 17,359 | 1.8% | Redis Cache (156 clusters) | High cluster count, many command-specific metrics |
| 3 | AWS/Kafka | 10,891 | 1.2% | MSK (3 clusters) | Per-topic/consumer-group lag metrics |
| 4 | AWS/EBS | 4,985 | 0.5% | Block Storage (~332 volumes) | Standard per-volume metrics |
| 5 | AWS/EC2 | 4,256 | 0.5% | Compute (233 instances) | Standard per-instance metrics |
| 6 | AWS/ES | 2,011 | 0.2% | OpenSearch (4 domains) | Comprehensive per-node metrics |
| 7 | AWS/RDS | 1,769 | 0.2% | MySQL/PostgreSQL (74 instances) | Standard DB metrics |
| 8 | AWS/DocDB | 1,097 | 0.1% | DocumentDB (3 clusters) | Standard DocDB metrics |
| 9 | AWS/NetworkELB | 953 | 0.1% | Network Load Balancers | Standard NLB metrics |
| 10 | AWS/Usage | 812 | 0.1% | AWS Usage Tracking | API call counts per service |

---

## Namespace Inventory (All {len(NAMESPACE_DATA)} Namespaces)

| Namespace | Metrics | Category | Type |
|-----------|---------|----------|------|
"""
    for ns, count in sorted(NAMESPACE_DATA.items(), key=lambda x: -x[1]):
        cat = CATEGORY_MAP.get(ns, "Other")
        ns_type = "AWS Native" if ns.startswith("AWS/") else "Custom"
        flag = " **<<< CRITICAL**" if count > 100000 else (" *HIGH*" if count > 5000 else "")
        report += f"| {ns} | {count:,}{flag} | {cat} | {ns_type} |\n"

    report += f"""
---

## Detailed Analysis by System Category

### 1. EMR Serverless (895,562 metrics -- 95.2% of all metrics)

**This is the single biggest cost optimization opportunity.**

- **8 per-worker metrics** (CPU, Memory, Storage allocated/used, StorageRead/WriteBytes) each have **~112,000 unique dimension combinations**
- These dimensions represent individual workers from historical EMR Serverless job runs
- EMR Serverless automatically publishes per-worker metrics that persist in CloudWatch even after jobs complete
- Application-level aggregate metrics (CPUAllocated, RunningWorkerCount, etc.) only account for ~36 metrics

**Recommendation:** Investigate EMR Serverless applications and completed job runs. Consider:
1. Cleaning up old EMR Serverless applications to remove retained worker metrics
2. Configuring CloudWatch metric stream filters to exclude per-worker dimensions
3. Using application-level aggregate metrics instead of per-worker tracking

### 2. ElastiCache / Redis (17,359 metrics)

- **156 Redis clusters** (129 x cache.t4g.micro, 15 x cache.t4g.small, 6 x cache.t3.micro, 4 x cache.t4g.medium, 2 x cache.m6g.large)
- 42 standard metrics per cluster + additional command-type metrics
- Command-specific metrics (HashBasedCmds, SetBasedCmds, ListBasedCmds, etc.) add cardinality
- All are AWS-published standard metrics -- no additional custom metric cost

### 3. Kafka/MSK (10,891 metrics)

- **3 MSK clusters:** iprod-kafka-architecture-cluster, iprod-kafka-base-cluster, iprod-kafka-business-cluster
- **OffsetLag** alone has 1,920 dimension combinations (per-topic, per-consumer-group, per-partition)
- EstimatedTimeLag: 1,918 dimensions
- Per-broker metrics: ~80 types x 9 brokers (3 per cluster)
- **Review:** Clean up stale consumer groups to reduce lag metric cardinality

### 4. EC2 Compute (4,256 metrics)

- **233 running instances**, all in us-east-1
- 18 standard metric types per instance
- **8 instances have detailed monitoring enabled** (1-minute resolution) -- mostly unnamed, likely EKS nodes
- 225 instances use basic monitoring (5-minute resolution)
- Detailed monitoring adds ~$2.10/instance/month = $16.80/month for 8 instances

### 5. OpenSearch/Elasticsearch (2,011 metrics)

- **4 domains:** luckycommon, luckylfe-log, luckyur-log, luckyus-opensearch-dify
- 52+ per-node metrics + 100+ cluster-level metrics
- Comprehensive metric coverage including KNN, ML, alerting, dashboards
- All AWS-published standard metrics

### 6. RDS MySQL/PostgreSQL (1,769 metrics)

- **74 DB instances:** 62 MySQL, 3 PostgreSQL (+ 9 DocDB counted separately)
- Instance classes: 41 x db.t4g.micro, 17 x db.t4g.medium, 6 x db.t3.medium
- Standard RDS metrics + MySQL-specific (BinLog, LVM) + PostgreSQL-specific (Checkpoint, TransactionLogs)
- Performance Insights enabled on 6 instances (separate cost)

### 7. DocumentDB (1,097 metrics)

- **3 clusters:** docdb-devops, docdb-gia, docdb-iot (9 instances total)
- 54+ standard metric types per instance/cluster
- Standard AWS-published metrics

---

## CloudWatch Alarms (28 total)

| State | Count |
|-------|-------|
| OK | 25 |
| INSUFFICIENT_DATA | 3 |
| ALARM | 0 |

### Alarm Breakdown

| Alarm Group | Namespace | Count | State | Notes |
|-------------|-----------|-------|-------|-------|
| Kafka partition/disk/offline | AWS/Kafka | 9 | OK | Critical MSK monitoring - keep |
| CloudFront domain 5xx alerts | AWS/CloudFront | 7 | 5 OK, 2 INSUF | Review 2 in INSUFFICIENT_DATA |
| API Gateway img-transform | AWS/ApiGateway | 3 | OK | Image transform pipeline |
| Lambda img-transform | AWS/Lambda | 3 | OK | Image transform pipeline |
| CDN img-transform | AWS/CloudFront | 3 | OK | Image transform pipeline |
| EMR Classic | AWS/ElasticMapReduce | 2 | INSUFFICIENT_DATA | **Review - clusters may be decommissioned** |
| ES disk space | Composite | 1 | OK | OpenSearch monitoring |

---

## EC2 Detailed Monitoring (8 instances)

| Instance ID | Name | Type | Cost Impact |
|-------------|------|------|-------------|
| i-0d9b1e751b082509e | unnamed | m6a.xlarge | $2.10/mo |
| i-0b7168740da06034a | unnamed | m6a.large | $2.10/mo |
| i-031550c05c4eae3e9 | unnamed | m6a.large | $2.10/mo |
| i-0f5b8098e2098d6ec | unnamed | r6i.2xlarge | $2.10/mo |
| i-0d812309d4071771c | unnamed | r6i.2xlarge | $2.10/mo |
| i-0047137d18fc77c35 | velodb-agent | c5.large | $2.10/mo |
| i-0a1e3602ff0d0cccf | unnamed | m6a.xlarge | $2.10/mo |
| i-0418d2213e6688dc0 | unnamed | m6a.large | $2.10/mo |

**Total detailed monitoring cost: ~$16.80/month**

---

## Cost Trend (3-Month)

| Usage Type | Nov 2025 | Dec 2025 | Jan 2026 | Trend |
|------------|----------|----------|----------|-------|
| MetricMonitorUsage | $1,920.66 | $1,891.83 | $1,872.68 | Slight decrease |
| Requests | $982.31 | $1,503.68 | $1,507.72 | Stable (after Nov jump) |
| GMD-Metrics | $3.46 | $3.60 | **$231.91** | **64x SPIKE** |
| InternetMonitor | $42.80 | $43.16 | $43.97 | Stable |
| Other | $10.88 | $8.51 | $10.50 | Stable |
| **TOTAL** | **$2,960.11** | **$3,450.78** | **$3,666.78** | **+6.3% MoM** |

---

## Recommendations & Action Plan

### Quick Wins (< 1 week, ~$237-274/month savings)

| # | Action | Savings | Effort |
|---|--------|---------|--------|
| 1 | **Investigate GMD-Metrics 64x spike** - Identify source of excessive GetMetricData queries (new Grafana dashboards, monitoring tools, automated scripts). Reduce query frequency or scope. | $200-230/mo | 1-3 days |
| 2 | **Review Internet Monitor** - Evaluate if $44/mo Internet Monitor service is actively used or can be replaced by CloudFront + Route53 native monitoring | $37-44/mo | 1 day |
| 3 | **Clean up EMR Classic alarms** - 2 alarms in INSUFFICIENT_DATA, likely from decommissioned clusters | ~$0.20/mo | 1 hour |
| 4 | **Review CloudFront alarms** - 2 alarms in INSUFFICIENT_DATA for potentially inactive distributions | ~$0.20/mo | 1 hour |

### Medium-term (1-2 weeks, ~$805-1,215/month additional)

| # | Action | Savings | Effort |
|---|--------|---------|--------|
| 5 | **EMR Serverless metric cleanup** - Address 895K+ per-worker metrics by cleaning up old applications or configuring metric suppression | $800-1,200/mo | 1-2 weeks |
| 6 | **Kafka consumer group audit** - Clean up stale consumer groups generating 1,920+ OffsetLag dimensions | $5-15/mo | 1 week |
| 7 | **EC2 detailed monitoring review** - Disable detailed monitoring on 8 instances if 1-minute resolution isn't needed | $16.80/mo | 1 hour |

### Longer-term Optimization

| # | Action | Savings | Effort |
|---|--------|---------|--------|
| 8 | **ElastiCache command metrics review** - Evaluate if all per-command-type metrics across 156 clusters are monitored | $0-5/mo | Ongoing |
| 9 | **RDS Performance Insights audit** - Verify PI is enabled only where needed | $0-10/mo | 1 day |
| 10 | **WorkMail ActionDenied investigation** - 79 denied action dimensions may indicate misconfiguration | $0 | 1 day |

### Total Savings Projection

| Timeframe | Monthly Savings | Annual Savings |
|-----------|----------------|----------------|
| Quick Wins | $237-274 | $2,844-3,288 |
| + Medium-term | $1,042-1,489 | $12,504-17,868 |
| Current Spend | $3,667/month | $44,004/year |
| Optimized Spend | ~$2,178-2,625/month | ~$26,136-31,500/year |
| **Savings %** | **28-41%** | **28-41%** |

---

## Infrastructure Summary

| Category | Resource | Count | Notes |
|----------|----------|-------|-------|
| Compute | EC2 Instances | 233 | All running |
| Compute | Lambda Functions | 13 | Image transform + others |
| Database | RDS MySQL | 62 instances | Primarily t4g.micro/medium |
| Database | RDS PostgreSQL | 3 instances | |
| Database | DocumentDB | 3 clusters / 9 instances | devops, gia, iot |
| Cache | ElastiCache Redis | 156 clusters | Mostly t4g.micro |
| Messaging | MSK Kafka | 3 clusters / 9 brokers | architecture, base, business |
| Search | OpenSearch | 4 domains | luckycommon, logs, dify |
| Storage | EBS Volumes | ~332 | |
| Storage | S3 Buckets | 34 | |
| CDN | CloudFront | 11 distributions | |
| Data Processing | EMR Serverless | 2+ applications | Source of 895K metrics |
| Networking | NLB | 50+ | |
| Networking | ALB | 4 | |
| Networking | NAT Gateway | 4 | |
| Networking | Transit Gateway | 9+ attachments | |

---

*Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M UTC')} by CloudWatch Cost Optimization Audit Tool*
*Luckin Coffee North America - DevOps DBA Team*
"""
    output_path = "/app/cloudwatch-cost-audit-2026-02/cloudwatch_cost_audit_luckin_2026-02.md"
    with open(output_path, 'w') as f:
        f.write(report)
    print(f"Markdown saved to {output_path}")
    return output_path


if __name__ == "__main__":
    xlsx_path = create_xlsx_report()
    md_path = create_markdown_report()
    print(f"\nReports generated:")
    print(f"  XLSX: {xlsx_path}")
    print(f"  Markdown: {md_path}")
